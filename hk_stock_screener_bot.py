#!/usr/bin/env python3
"""
HK Stock Screening & Multi-Channel Alert Bot (Telegram + WhatsApp)
Production-grade automated Hong Kong stock screening with Excel reports and alerts.
"""

# For Jupyter users: uncomment and run to install dependencies
# !pip install yfinance pandas numpy matplotlib mplfinance openpyxl python-telegram-bot twilio ta requests nest_asyncio schedule

# ================================================================
# IMPORTS
# ================================================================

# Standard library
import asyncio
import gc
import logging
import os
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

# Third party
import numpy as np
import pandas as pd

try:
    import nest_asyncio
    nest_asyncio.apply()
except ImportError:
    pass

# ================================================================
# CONFIGURATION
# ================================================================

# --- Paths ---
INPUT_EXCEL_PATH = "aastocks_export.xlsx"
OUTPUT_EXCEL_PATH = "stock_screening_results.xlsx"
OUTPUT_CHART_PATH = "top_performers_chart.png"
OUTPUT_CANDLESTICK_DIR = "candlestick_charts"
DUPLICATE_MARKER_FILE = ".last_sent_date.txt"
LOG_FILE_PATH = "bot_run_log.txt"

# --- Data ---
DATA_PERIOD = "1y"  # yfinance period: 1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max
TOP_N_STOCKS = 10   # Number of top/bottom performers for reports and charts
TOP_N_EXCEL = 20    # Rows for "Top Performers" and "Bottom Performers" sheets

# --- Technical Indicator Parameters ---
ADX_PERIOD = 14
RSI_PERIOD = 14
BOLLINGER_STD = 2

# --- Alert Thresholds ---
VOLUME_SPIKE_THRESHOLD = 2.0   # Flag if today's volume >= 2x the 20-day avg
HIGH_ADX_THRESHOLD = 25        # ADX above this = strong trend
RSI_OVERBOUGHT = 70
RSI_OVERSOLD = 30
ZSCORE_EXTREME_HIGH = 2.0
ZSCORE_EXTREME_LOW = -2.0

# --- Rate Limiting & Network ---
DOWNLOAD_DELAY_SECONDS = 0.1
DOWNLOAD_TIMEOUT_SECONDS = 30
TELEGRAM_RETRY_ATTEMPTS = 2
TELEGRAM_RETRY_DELAY_SECONDS = 5
WHATSAPP_RETRY_ATTEMPTS = 2
WHATSAPP_RETRY_DELAY_SECONDS = 5
MIN_DATA_DAYS = 60

# --- Channels ---
TELEGRAM_ENABLED = True
TELEGRAM_BOT_TOKEN = ""  # Set your bot token
TELEGRAM_CHAT_ID = ""    # Set your chat ID

WHATSAPP_ENABLED = False
TWILIO_ACCOUNT_SID = ""
TWILIO_AUTH_TOKEN = ""
TWILIO_WHATSAPP_FROM = "whatsapp:+14155238886"  # Sandbox number
TWILIO_WHATSAPP_TO = "whatsapp:+852XXXXXXXX"

# --- Override duplicate check ---
FORCE_SEND = False

# --- Scheduling ---
SCHEDULER_ENABLED = False
SCHEDULER_RUN_TIME = "18:30"

# Fallback stock list when Excel is not found
FALLBACK_STOCKS = [
    "0001.HK", "0002.HK", "0003.HK", "0005.HK", "0006.HK",
    "0011.HK", "0012.HK", "0016.HK", "0017.HK", "0027.HK",
    "0066.HK", "0101.HK", "0175.HK", "0241.HK", "0267.HK",
    "0288.HK", "0388.HK", "0669.HK", "0700.HK", "0762.HK",
    "0823.HK", "0857.HK", "0883.HK", "0939.HK", "0941.HK",
    "1038.HK", "1044.HK", "1093.HK", "1177.HK", "1299.HK",
    "1398.HK", "1810.HK", "1928.HK", "2007.HK", "2018.HK",
    "2020.HK", "2269.HK", "2313.HK", "2318.HK", "2328.HK",
    "2382.HK", "2388.HK", "2628.HK", "2688.HK", "3328.HK",
    "3988.HK", "6098.HK", "6862.HK", "9618.HK", "9888.HK",
    "9988.HK", "9999.HK",
]

# ================================================================
# LOGGING SETUP
# ================================================================

logging.basicConfig(
    level=logging.DEBUG,
    format="[%(asctime)s] %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOG_FILE_PATH, encoding="utf-8"),
    ],
)
# Set root logger to INFO for console, DEBUG for file
logging.getLogger().handlers[0].setLevel(logging.INFO)
logging.getLogger().handlers[1].setLevel(logging.DEBUG)
logger = logging.getLogger(__name__)


# ================================================================
# SECTION: Input Data
# ================================================================

def load_stock_list(filepath: str) -> list[str]:
    """
    Load stock codes from an AASTOCKS Excel export.

    Reads Excel (tries openpyxl then xlrd), detects code column by header patterns,
    normalizes codes to 4 digits + .HK, deduplicates. Falls back to FALLBACK_STOCKS
    if file is missing or yields no valid codes.

    Args:
        filepath: Path to the Excel file.

    Returns:
        List of ticker strings (e.g. ["0005.HK", "0700.HK"]).
    """
    code_headers = [
        "code", "stock code", "ticker", "symbol",
        "代號", "股票代號", "编号",
    ]
    engines = ["openpyxl", None]
    if not os.path.isfile(filepath):
        logger.warning(f"File not found: {filepath}. Using fallback stock list.")
        logger.info(f"Loaded {len(FALLBACK_STOCKS)} stock codes from fallback")
        return FALLBACK_STOCKS.copy()

    for engine in engines:
        try:
            if engine:
                df = pd.read_excel(filepath, engine=engine, sheet_name=None)
            else:
                df = pd.read_excel(filepath, sheet_name=None)
        except Exception as e:
            logger.debug(f"Engine {engine} failed: {e}")
            continue

        if not isinstance(df, dict):
            df = {0: df}

        for sheet_name, sheet_df in df.items():
            if sheet_df is None or sheet_df.empty:
                continue
            sheet_df = sheet_df.astype(str)
            cols = [c.strip() for c in sheet_df.columns.astype(str)]
            code_col = None
            for h in code_headers:
                for i, c in enumerate(cols):
                    if h in c.lower():
                        code_col = i
                        break
                if code_col is not None:
                    break
            if code_col is None:
                code_col = 0
            raw = sheet_df.iloc[:, code_col].dropna().astype(str).str.strip()
            codes = []
            for v in raw:
                digits = re.sub(r"\D", "", v)
                if not digits:
                    continue
                padded = digits.zfill(4)
                if len(padded) > 4:
                    padded = padded[:4]
                ticker = padded + ".HK"
                codes.append(ticker)
            codes = list(dict.fromkeys(codes))
            if codes:
                logger.info(f"Loaded {len(codes)} stock codes from {filepath} (sheet: {sheet_name})")
                return codes
    logger.warning(f"No valid codes from {filepath}. Using fallback stock list.")
    logger.info(f"Loaded {len(FALLBACK_STOCKS)} stock codes from fallback")
    return FALLBACK_STOCKS.copy()


# ================================================================
# SECTION: Data Download
# ================================================================

def download_all_data(
    ticker_list: list[str], period: str
) -> tuple[dict[str, pd.DataFrame], pd.DataFrame]:
    """
    Download OHLCV history for each ticker and for benchmark ^HSI.

    Uses yfinance per-ticker (no batch) with delay and timeout. Skips tickers
    with insufficient rows (< MIN_DATA_DAYS). Returns stock dict and benchmark DataFrame.

    Args:
        ticker_list: List of yfinance ticker strings (e.g. 0005.HK).
        period: yfinance period string (e.g. 1y).

    Returns:
        (stock_data, benchmark_data). stock_data maps ticker -> DataFrame;
        benchmark_data is HSI OHLCV or empty DataFrame on failure.
    """
    import yfinance as yf

    stock_data = {}
    total = len(ticker_list)
    success = 0
    failed = 0
    skipped = 0

    for i, ticker in enumerate(ticker_list):
        ts = datetime.now().strftime("%H:%M:%S")
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period=period, timeout=DOWNLOAD_TIMEOUT_SECONDS)
            time.sleep(DOWNLOAD_DELAY_SECONDS)
            if hist is None or len(hist) < MIN_DATA_DAYS:
                logger.info(f"[{ts}] Downloading {i+1}/{total}: {ticker}... ✗ (Insufficient data: {len(hist) if hist is not None else 0} days)")
                skipped += 1
                continue
            stock_data[ticker] = hist
            success += 1
            logger.info(f"[{ts}] Downloading {i+1}/{total}: {ticker}... ✓ ({len(hist)} days)")
        except Exception as e:
            logger.info(f"[{ts}] Downloading {i+1}/{total}: {ticker}... ✗ ({e})")
            failed += 1

    benchmark_data = pd.DataFrame()
    try:
        b = yf.Ticker("^HSI")
        benchmark_data = b.history(period=period, timeout=DOWNLOAD_TIMEOUT_SECONDS)
        if benchmark_data is not None and not benchmark_data.empty:
            logger.info("Benchmark (^HSI): ✓")
        else:
            logger.warning("Benchmark (^HSI): No data")
            benchmark_data = pd.DataFrame()
    except Exception as e:
        logger.warning(f"Benchmark (^HSI): ✗ ({e})")

    logger.info(
        f"Download Summary:\n  Total attempted: {total}\n  Successful: {success}\n  Failed: {failed}\n  Skipped (insufficient data): {skipped}\n  Benchmark (^HSI): {'✓' if not benchmark_data.empty else '✗'}"
    )
    return stock_data, benchmark_data


# ================================================================
# SECTION: Technical Indicators
# ================================================================

def _wilder_smooth(series: pd.Series, period: int) -> pd.Series:
    """First value = sum of first period; then recursive Wilder smoothing."""
    out = pd.Series(index=series.index, dtype=float)
    out.iloc[period - 1] = series.iloc[:period].sum()
    for i in range(period, len(series)):
        out.iloc[i] = out.iloc[i - 1] - (out.iloc[i - 1] / period) + series.iloc[i]
    return out


def calculate_indicators(ticker: str, df: pd.DataFrame) -> dict:
    """
    Compute all technical indicators and metrics for one stock.

    Handles insufficient data by returning np.nan for affected metrics.
    Returns a flat dict with Ticker and all indicator keys.

    Args:
        ticker: Stock ticker string (e.g. 0700.HK).
        df: OHLCV DataFrame with DatetimeIndex.

    Returns:
        Single flat dictionary of metrics (floats/str, rounded as specified).
    """
    out = {"Ticker": ticker}
    close = df["Close"]
    high = df["High"]
    low = df["Low"]
    n = len(close)
    if n < 2:
        for k in [
            "1d_return", "5d_return", "20d_return", "60d_return", "ytd_return",
            "gmma_trend", "gmma_trend_duration", "gmma_spread", "zscore",
            "adx", "di_plus", "di_minus", "rsi",
            "macd_line", "macd_signal", "macd_histogram", "macd_crossover",
            "bb_percent_b", "bb_bandwidth",
            "avg_volume_20d", "volume_ratio", "volume_trend",
            "latest_close", "high_52w", "low_52w", "dist_from_52w_high", "dist_from_52w_low",
            "sma_50", "sma_200", "ma_cross",
        ]:
            out[k] = np.nan if k != "gmma_trend" and k != "volume_trend" and k != "macd_crossover" and k != "ma_cross" else ("Unknown" if k in ("gmma_trend", "volume_trend") else "None")
        return out

    # --- Returns ---
    out["1d_return"] = round((close.iloc[-1] / close.iloc[-2] - 1) * 100, 2) if n >= 2 else np.nan
    out["5d_return"] = round((close.iloc[-1] / close.iloc[-6] - 1) * 100, 2) if n >= 6 else np.nan
    out["20d_return"] = round((close.iloc[-1] / close.iloc[-21] - 1) * 100, 2) if n >= 21 else np.nan
    out["60d_return"] = round((close.iloc[-1] / close.iloc[-61] - 1) * 100, 2) if n >= 61 else np.nan
    try:
        year_start = pd.Timestamp(datetime.now().year, 1, 1)
        mask = close.index >= year_start
        if mask.any():
            first_close = close.loc[mask].iloc[0]
            out["ytd_return"] = round((close.iloc[-1] / first_close - 1) * 100, 2)
        else:
            out["ytd_return"] = np.nan
    except Exception:
        out["ytd_return"] = np.nan

    # --- GMMA ---
    short_periods = [3, 5, 8, 10, 12, 15]
    long_periods = [30, 35, 40, 45, 50, 60]
    short_emas = [close.ewm(span=p, adjust=False).mean() for p in short_periods]
    long_emas = [close.ewm(span=p, adjust=False).mean() for p in long_periods]
    short_avg = sum(short_emas) / 6
    long_avg = sum(long_emas) / 6
    cur_short = short_avg.iloc[-1]
    cur_long = long_avg.iloc[-1]
    out["gmma_trend"] = "Bullish" if cur_short > cur_long else "Bearish"
    out["gmma_spread"] = round((cur_short - cur_long) / cur_long * 100, 2) if cur_long else np.nan
    duration = 0
    current_bull = cur_short > cur_long
    for j in range(len(short_avg) - 1, -1, -1):
        if (short_avg.iloc[j] > long_avg.iloc[j]) == current_bull:
            duration += 1
        else:
            break
    out["gmma_trend_duration"] = duration

    # --- Z-Score ---
    sma20 = close.rolling(20).mean().iloc[-1]
    std20 = close.rolling(20).std().iloc[-1]
    if std20 and std20 > 0:
        out["zscore"] = round((close.iloc[-1] - sma20) / std20, 2)
    else:
        out["zscore"] = np.nan

    # --- ADX / DI+ / DI- ---
    try:
        from ta.trend import ADXIndicator
        adx_ind = ADXIndicator(high=high, low=low, close=close, window=ADX_PERIOD)
        out["adx"] = round(adx_ind.adx().iloc[-1], 2)
        out["di_plus"] = round(adx_ind.adx_pos().iloc[-1], 2)
        out["di_minus"] = round(adx_ind.adx_neg().iloc[-1], 2)
    except Exception:
        prev_close = close.shift(1)
        prev_high = high.shift(1)
        prev_low = low.shift(1)
        tr = pd.concat([
            high - low,
            (high - prev_close).abs(),
            (low - prev_close).abs(),
        ], axis=1).max(axis=1)
        plus_dm = np.where((high - prev_high > prev_low - low) & (high > prev_high), high - prev_high, 0.0)
        minus_dm = np.where((prev_low - low > high - prev_high) & (low < prev_low), prev_low - low, 0.0)
        plus_dm = pd.Series(plus_dm, index=close.index)
        minus_dm = pd.Series(minus_dm, index=close.index)
        tr_s = _wilder_smooth(tr, ADX_PERIOD)
        plus_s = _wilder_smooth(plus_dm, ADX_PERIOD)
        minus_s = _wilder_smooth(minus_dm, ADX_PERIOD)
        di_plus = 100 * plus_s / tr_s.replace(0, np.nan)
        di_minus = 100 * minus_s / tr_s.replace(0, np.nan)
        dx = 100 * (di_plus - di_minus).abs() / (di_plus + di_minus).replace(0, np.nan)
        adx = _wilder_smooth(dx.fillna(0), ADX_PERIOD)
        out["adx"] = round(adx.iloc[-1], 2)
        out["di_plus"] = round(di_plus.iloc[-1], 2)
        out["di_minus"] = round(di_minus.iloc[-1], 2)

    # --- RSI ---
    try:
        from ta.momentum import RSIIndicator
        rsi_ser = RSIIndicator(close=close, window=RSI_PERIOD).rsi()
        out["rsi"] = round(rsi_ser.iloc[-1], 2) if not pd.isna(rsi_ser.iloc[-1]) else np.nan
    except Exception:
        delta = close.diff()
        gain = delta.where(delta > 0, 0.0)
        loss = (-delta).where(delta < 0, 0.0)
        if n >= RSI_PERIOD:
            avg_gain = gain.rolling(RSI_PERIOD).mean()
            avg_loss = loss.rolling(RSI_PERIOD).mean()
            rs = avg_gain / avg_loss.replace(0, np.nan)
            rsi = 100 - (100 / (1 + rs))
            out["rsi"] = round(rsi.iloc[-1], 2) if not pd.isna(rsi.iloc[-1]) else np.nan
        else:
            out["rsi"] = np.nan

    # --- MACD ---
    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    macd_signal = macd_line.ewm(span=9, adjust=False).mean()
    macd_hist = macd_line - macd_signal
    out["macd_line"] = round(macd_line.iloc[-1], 4)
    out["macd_signal"] = round(macd_signal.iloc[-1], 4)
    out["macd_histogram"] = round(macd_hist.iloc[-1], 4)
    crossover = "None"
    if n >= 4:
        for k in range(1, 4):
            if macd_line.iloc[-k - 1] <= macd_signal.iloc[-k - 1] and macd_line.iloc[-k] > macd_signal.iloc[-k]:
                crossover = "Bullish Cross"
                break
            if macd_line.iloc[-k - 1] >= macd_signal.iloc[-k - 1] and macd_line.iloc[-k] < macd_signal.iloc[-k]:
                crossover = "Bearish Cross"
                break
    out["macd_crossover"] = crossover

    # --- Bollinger ---
    mid = close.rolling(20).mean()
    std = close.rolling(20).std()
    upper = mid + BOLLINGER_STD * std
    lower = mid - BOLLINGER_STD * std
    width = (upper - lower) / mid.replace(0, np.nan) * 100
    pct_b = (close - lower) / (upper - lower).replace(0, np.nan)
    out["bb_percent_b"] = round(pct_b.iloc[-1], 4) if not pd.isna(pct_b.iloc[-1]) else np.nan
    out["bb_bandwidth"] = round(width.iloc[-1], 2) if not pd.isna(width.iloc[-1]) else np.nan

    # --- Volume ---
    vol = df["Volume"] if "Volume" in df.columns else pd.Series(0, index=df.index)
    avg_vol_20 = vol.rolling(20).mean().iloc[-1]
    out["avg_volume_20d"] = round(avg_vol_20, 0) if not pd.isna(avg_vol_20) else np.nan
    last_vol = vol.iloc[-1]
    out["volume_ratio"] = round(last_vol / avg_vol_20, 2) if avg_vol_20 and avg_vol_20 > 0 else np.nan
    avg_5 = vol.rolling(5).mean().iloc[-1]
    if not pd.isna(avg_vol_20) and avg_vol_20 > 0 and not pd.isna(avg_5):
        if avg_5 > avg_vol_20 * 1.5:
            out["volume_trend"] = "Increasing"
        elif avg_5 < avg_vol_20 * 0.5:
            out["volume_trend"] = "Decreasing"
        else:
            out["volume_trend"] = "Normal"
    else:
        out["volume_trend"] = "Normal"

    # --- Price levels ---
    out["latest_close"] = round(close.iloc[-1], 3)
    look = min(252, n)
    high_52 = close.iloc[-look:].max()
    low_52 = close.iloc[-look:].min()
    out["high_52w"] = round(high_52, 3)
    out["low_52w"] = round(low_52, 3)
    out["dist_from_52w_high"] = round((close.iloc[-1] - high_52) / high_52 * 100, 2) if high_52 else np.nan
    out["dist_from_52w_low"] = round((close.iloc[-1] - low_52) / low_52 * 100, 2) if low_52 else np.nan
    sma50 = close.rolling(50).mean().iloc[-1]
    sma200 = close.rolling(200).mean().iloc[-1]
    out["sma_50"] = round(sma50, 3) if not pd.isna(sma50) else np.nan
    out["sma_200"] = round(sma200, 3) if not pd.isna(sma200) else np.nan
    ma_cross = "None"
    if n >= 55:
        s50 = close.rolling(50).mean()
        s200 = close.rolling(200).mean()
        if s50.iloc[-6] < s200.iloc[-6] and s50.iloc[-1] > s200.iloc[-1]:
            ma_cross = "Golden Cross"
        elif s50.iloc[-6] > s200.iloc[-6] and s50.iloc[-1] < s200.iloc[-1]:
            ma_cross = "Death Cross"
    out["ma_cross"] = ma_cross

    return out


# ================================================================
# SECTION: Alert Flags
# ================================================================

def generate_alerts(row: dict) -> list[str]:
    """
    Generate list of human-readable alert strings from one stock's metrics.

    Uses CONFIG thresholds. Returns list of strings (e.g. "🔥 Volume Spike: 2.3x average").
    """
    alerts = []
    vr = row.get("volume_ratio")
    if vr is not None and vr != np.nan and vr >= VOLUME_SPIKE_THRESHOLD:
        alerts.append(f"🔥 Volume Spike: {vr:.1f}x average")
    adx, dp, dm = row.get("adx"), row.get("di_plus"), row.get("di_minus")
    if adx is not None and adx != np.nan and adx > HIGH_ADX_THRESHOLD:
        if dp is not None and dm is not None and dp > dm:
            alerts.append(f"🚀 Strong Uptrend (ADX: {adx:.1f})")
        elif dp is not None and dm is not None and dm > dp:
            alerts.append(f"📉 Strong Downtrend (ADX: {adx:.1f})")
    rsi = row.get("rsi")
    if rsi is not None and rsi != np.nan:
        if rsi > RSI_OVERBOUGHT:
            alerts.append(f"⚠️ RSI Overbought: {rsi:.1f}")
        if rsi < RSI_OVERSOLD:
            alerts.append(f"💰 RSI Oversold: {rsi:.1f}")
    if row.get("macd_crossover") == "Bullish Cross":
        alerts.append("📊 MACD Bullish Crossover")
    if row.get("macd_crossover") == "Bearish Cross":
        alerts.append("📊 MACD Bearish Crossover")
    if row.get("ma_cross") == "Golden Cross":
        alerts.append("🌟 SMA 50/200 Golden Cross")
    if row.get("ma_cross") == "Death Cross":
        alerts.append("💀 SMA 50/200 Death Cross")
    z = row.get("zscore")
    if z is not None and z != np.nan:
        if z > ZSCORE_EXTREME_HIGH:
            alerts.append(f"📈 Z-Score Extreme: {z:.2f} (potentially overextended)")
        if z < ZSCORE_EXTREME_LOW:
            alerts.append(f"📉 Z-Score Extreme: {z:.2f} (potentially oversold)")
    d_high = row.get("dist_from_52w_high")
    if d_high is not None and d_high != np.nan and d_high > -2.0:
        alerts.append("🔔 Within 2% of 52-Week High")
    d_low = row.get("dist_from_52w_low")
    if d_low is not None and d_low != np.nan and d_low < 2.0:
        alerts.append("🔔 Within 2% of 52-Week Low")
    if row.get("gmma_trend") == "Bullish":
        dur = row.get("gmma_trend_duration")
        if dur is not None and dur != np.nan and dur > 20:
            alerts.append(f"🐂 Sustained Bullish GMMA ({int(dur)} days)")
    bb = row.get("bb_percent_b")
    if bb is not None and bb != np.nan:
        if bb > 1.0:
            alerts.append("🔴 Price above Upper Bollinger Band")
        if bb < 0.0:
            alerts.append("🟢 Price below Lower Bollinger Band")
    return alerts


# ================================================================
# SECTION: Screening & Ranking
# ================================================================

def screen_stocks(all_stock_data: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Compute indicators and alerts for all stocks and return one DataFrame.

    Sorts by 20-Day Return descending and adds Rank. Adds Alerts and Alert_Count.
    """
    rows = []
    total = len(all_stock_data)
    for i, (ticker, df) in enumerate(all_stock_data.items()):
        if (i + 1) % 50 == 0 or i == 0:
            logger.info(f"Calculating indicators: {i+1}/{total}...")
        try:
            metrics = calculate_indicators(ticker, df)
            alerts = generate_alerts(metrics)
            metrics["Alerts"] = " | ".join(alerts)
            metrics["Alert_Count"] = len(alerts)
            metrics["Stock Code"] = ticker.replace(".HK", "")
            rows.append(metrics)
        except Exception as e:
            logger.warning(f"Failed indicators for {ticker}: {e}")
    if not rows:
        return pd.DataFrame()
    results_df = pd.DataFrame(rows)
    results_df = results_df.sort_values("20d_return", ascending=False).reset_index(drop=True)
    results_df["Rank"] = range(1, len(results_df) + 1)
    return results_df


# ================================================================
# SECTION: Market Summary
# ================================================================

def compute_market_summary(results_df: pd.DataFrame) -> dict:
    """Compute market breadth and summary stats from screened results."""
    if results_df is None or results_df.empty:
        return {}
    total = len(results_df)
    bullish = (results_df["gmma_trend"] == "Bullish").sum()
    bearish = (results_df["gmma_trend"] == "Bearish").sum()
    bull_ratio = round(bullish / total * 100, 2) if total else 0
    avg_1d = results_df["1d_return"].mean()
    avg_5d = results_df["5d_return"].mean()
    avg_20d = results_df["20d_return"].mean()
    pos_20d = (results_df["20d_return"] > 0).sum()
    breadth_20d = round(pos_20d / total * 100, 2) if total else 0
    above_50 = (results_df["latest_close"] > results_df["sma_50"]).sum()
    above_200 = (results_df["latest_close"] > results_df["sma_200"]).sum()
    breadth_50 = round(above_50 / total * 100, 2) if total else 0
    breadth_200 = round(above_200 / total * 100, 2) if total else 0
    vol_spikes = (results_df["volume_ratio"] >= VOLUME_SPIKE_THRESHOLD).sum()
    avg_rsi = results_df["rsi"].mean()
    rsi_over = (results_df["rsi"] > 70).sum()
    rsi_under = (results_df["rsi"] < 30).sum()
    total_alerts = results_df["Alert_Count"].sum()
    all_alerts = []
    for a in results_df["Alerts"].dropna():
        all_alerts.extend([x.strip() for x in str(a).split("|") if x.strip()])
    most_common = Counter(all_alerts).most_common(1)
    most_common_alert = most_common[0][0] if most_common else "N/A"
    return {
        "total_stocks": total,
        "gmma_bullish": int(bullish),
        "gmma_bearish": int(bearish),
        "bullish_ratio_pct": bull_ratio,
        "avg_1d_return": round(avg_1d, 2) if not pd.isna(avg_1d) else 0,
        "avg_5d_return": round(avg_5d, 2) if not pd.isna(avg_5d) else 0,
        "avg_20d_return": round(avg_20d, 2) if not pd.isna(avg_20d) else 0,
        "breadth_20d_pct": breadth_20d,
        "breadth_above_sma50_pct": breadth_50,
        "breadth_above_sma200_pct": breadth_200,
        "volume_spikes_count": int(vol_spikes),
        "avg_rsi": round(avg_rsi, 2) if not pd.isna(avg_rsi) else 0,
        "rsi_overbought_count": int(rsi_over),
        "rsi_oversold_count": int(rsi_under),
        "total_alerts": int(total_alerts),
        "most_common_alert": most_common_alert,
    }


def print_market_summary(market_summary: dict) -> None:
    """Log market summary to console/file."""
    if not market_summary:
        return
    logger.info("Market Summary:")
    logger.info(f"  Total Stocks: {market_summary['total_stocks']}")
    logger.info(f"  GMMA: {market_summary['gmma_bullish']} Bullish / {market_summary['gmma_bearish']} Bearish ({market_summary['bullish_ratio_pct']}% bullish)")
    logger.info(f"  Avg Returns: 1D {market_summary['avg_1d_return']}% | 5D {market_summary['avg_5d_return']}% | 20D {market_summary['avg_20d_return']}%")
    logger.info(f"  Breadth 20D positive: {market_summary['breadth_20d_pct']}% | Above SMA50: {market_summary['breadth_above_sma50_pct']}% | Above SMA200: {market_summary['breadth_above_sma200_pct']}%")
    logger.info(f"  Volume Spikes: {market_summary['volume_spikes_count']} | Avg RSI: {market_summary['avg_rsi']} | Overbought: {market_summary['rsi_overbought_count']} | Oversold: {market_summary['rsi_oversold_count']}")
    logger.info(f"  Total Alerts: {market_summary['total_alerts']} | Most common: {market_summary['most_common_alert']}")


# ================================================================
# SECTION: Excel Report
# ================================================================

def save_excel_report(
    results_df: pd.DataFrame, market_summary: dict, filepath: str
) -> str:
    """
    Save multi-sheet Excel workbook with formatted tables and styling.

    Sheets: Full Results, Top Performers, Bottom Performers, Bullish GMMA,
    High ADX Uptrend, RSI Extremes, Volume Spikes, MACD Crossovers,
    Alert Summary, Market Summary. Returns filepath of saved file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if results_df is None or results_df.empty:
        pd.DataFrame({"Message": ["No data"]}).to_excel(filepath, index=False)
        return filepath

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Full Results
        results_df.to_excel(writer, sheet_name="Full Results", index=False)
        # Top / Bottom
        top = results_df.head(TOP_N_EXCEL)
        bottom = results_df.tail(TOP_N_EXCEL)
        top.to_excel(writer, sheet_name="Top Performers", index=False)
        bottom.to_excel(writer, sheet_name="Bottom Performers", index=False)
        # Bullish GMMA
        bull = results_df[results_df["gmma_trend"] == "Bullish"].sort_values("gmma_trend_duration", ascending=False)
        bull.to_excel(writer, sheet_name="Bullish GMMA", index=False)
        # High ADX Uptrend
        adx_up = results_df[
            (results_df["adx"] > HIGH_ADX_THRESHOLD) & (results_df["di_plus"] > results_df["di_minus"])
        ].sort_values("adx", ascending=False)
        adx_up.to_excel(writer, sheet_name="High ADX Uptrend", index=False)
        # RSI Extremes
        rsi_ext = results_df[(results_df["rsi"] > 70) | (results_df["rsi"] < 30)].sort_values("rsi", ascending=False)
        rsi_ext.to_excel(writer, sheet_name="RSI Extremes", index=False)
        # Volume Spikes
        vol_sp = results_df[results_df["volume_ratio"] >= VOLUME_SPIKE_THRESHOLD].sort_values("volume_ratio", ascending=False)
        vol_sp.to_excel(writer, sheet_name="Volume Spikes", index=False)
        # MACD Crossovers
        macd_c = results_df[results_df["macd_crossover"] != "None"]
        macd_c.to_excel(writer, sheet_name="MACD Crossovers", index=False)
        # Alert Summary
        alert_df = results_df[results_df["Alert_Count"] > 0].sort_values("Alert_Count", ascending=False)
        alert_df.to_excel(writer, sheet_name="Alert Summary", index=False)
        # Market Summary
        ms_df = pd.DataFrame(list(market_summary.items()), columns=["Metric", "Value"])
        ms_df.to_excel(writer, sheet_name="Market Summary", index=False)

    wb = openpyxl.load_workbook(filepath)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    return_cols = ["1d_return", "5d_return", "20d_return", "60d_return", "ytd_return", "dist_from_52w_high", "dist_from_52w_low", "gmma_spread", "bb_bandwidth"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            max_len = 0
            for r in range(1, min(ws.max_row + 1, 500)):
                cell = ws.cell(row=r, column=c)
                if r == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                val = cell.value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 40))
            ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 40)
        if sheet_name == "Market Summary":
            continue
        try:
            hdr = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
            for r in range(2, ws.max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if c <= len(hdr) and hdr[c - 1] in return_cols:
                        try:
                            v = cell.value
                            if v is not None and isinstance(v, (int, float)):
                                if v > 0:
                                    cell.fill = green_fill
                                elif v < 0:
                                    cell.fill = red_fill
                        except Exception:
                            pass
                    if c <= len(hdr) and hdr[c - 1] == "Alerts" and cell.value:
                        cell.fill = yellow_fill
        except Exception:
            pass
        ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    return filepath


# ================================================================
# SECTION: Performance Chart
# ================================================================

def plot_top_performers(
    all_stock_data: dict[str, pd.DataFrame],
    top_tickers: list[str],
    benchmark_data: pd.DataFrame,
    filepath: str,
) -> None:
    """
    Plot normalized price performance and relative strength vs HSI.
    Saves figure to filepath as PNG. Closes figure after save.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if not top_tickers or not all_stock_data:
        return
    today = datetime.now().strftime("%Y-%m-%d")
    fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True, gridspec_kw={"height_ratios": [0.7, 0.3]}, figsize=(16, 10))
    colors = plt.cm.tab10(np.linspace(0, 1, len(top_tickers)))
    hsi_norm = None
    if benchmark_data is not None and not benchmark_data.empty and "Close" in benchmark_data.columns:
        hsi_close = benchmark_data["Close"]
        hsi_norm = (hsi_close / hsi_close.iloc[0]) * 100
        hsi_ret = (hsi_close.iloc[-1] / hsi_close.iloc[0] - 1) * 100
        ax1.plot(hsi_norm.index, hsi_norm.values, "k--", linewidth=2.5, label=f"^HSI ({hsi_ret:+.1f}%)")
    ax1.axhline(100, color="gray", linestyle="--", alpha=0.7)
    for idx, ticker in enumerate(top_tickers):
        if ticker not in all_stock_data:
            continue
        df = all_stock_data[ticker]
        close = df["Close"]
        norm = (close / close.iloc[0]) * 100
        ret = (close.iloc[-1] / close.iloc[0] - 1) * 100
        ax1.plot(norm.index, norm.values, color=colors[idx], label=f"{ticker} ({ret:+.1f}%)")
    ax1.set_ylabel("Normalized Price (Base = 100)")
    ax1.set_title(f"Top {len(top_tickers)} HK Stock Performers vs Hang Seng Index — {today}")
    ax1.legend(loc="best", fontsize=8)
    ax1.grid(alpha=0.3)
    if hsi_norm is not None:
        for idx, ticker in enumerate(top_tickers):
            if ticker not in all_stock_data:
                continue
            df = all_stock_data[ticker]
            close = df["Close"]
            norm = (close / close.iloc[0]) * 100
            common_idx = norm.index.intersection(hsi_norm.index)
            if len(common_idx) < 2:
                continue
            rel = (norm.reindex(common_idx).ffill().bfill() / hsi_norm.reindex(common_idx).ffill().bfill()) * 100
            ax2.plot(rel.index, rel.values, color=colors[idx], label=ticker)
        ax2.axhline(100, color="gray", linestyle="--", alpha=0.7)
    ax2.set_ylabel("Relative Strength vs HSI (Base = 100)")
    ax2.grid(alpha=0.3)
    ax2.legend(loc="best", fontsize=8)
    plt.tight_layout()
    plt.savefig(filepath, dpi=150, bbox_inches="tight")
    plt.close()


# ================================================================
# SECTION: Candlestick Charts
# ================================================================

def plot_candlestick_charts(
    all_stock_data: dict[str, pd.DataFrame],
    top_tickers: list[str],
    output_dir: str,
) -> list[str]:
    """
    Generate candlestick charts for top tickers (last 60 days, volume, SMA, BB).
    Returns list of saved image file paths.
    """
    os.makedirs(output_dir, exist_ok=True)
    paths = []
    try:
        import mplfinance as mpf
    except ImportError:
        logger.warning("mplfinance not installed. Skipping candlestick chart generation.")
        return paths
    for ticker in top_tickers:
        if ticker not in all_stock_data:
            continue
        try:
            df = all_stock_data[ticker].tail(60).copy()
            if len(df) < 20:
                continue
            df.columns = [c.lower() for c in df.columns]
            df["SMA20"] = df["close"].rolling(20).mean()
            df["SMA50"] = df["close"].rolling(50).mean()
            mid = df["close"].rolling(20).mean()
            std = df["close"].rolling(20).std()
            df["BB_upper"] = mid + BOLLINGER_STD * std
            df["BB_lower"] = mid - BOLLINGER_STD * std
            add_plot = [
                mpf.make_addplot(df["SMA20"], color="blue"),
                mpf.make_addplot(df["SMA50"], color="orange"),
                mpf.make_addplot(df["BB_upper"], color="gray", linestyle="--"),
                mpf.make_addplot(df["BB_lower"], color="gray", linestyle="--"),
            ]
            out_path = os.path.join(output_dir, f"{ticker}_candlestick.png")
            mpf.plot(
                df,
                type="candle",
                volume=True,
                addplot=add_plot,
                title=f"{ticker} — Last 60 Trading Days",
                style="yahoo",
                savefig=dict(fname=out_path, dpi=150),
            )
            paths.append(out_path)
        except Exception as e:
            logger.warning(f"Candlestick failed for {ticker}: {e}")
    return paths


# ================================================================
# SECTION: Telegram Integration
# ================================================================
"""
TELEGRAM BOT SETUP INSTRUCTIONS:

1. Open Telegram and search for @BotFather
2. Send /newbot and follow the prompts to create your bot
3. BotFather will give you an API token — copy it to TELEGRAM_BOT_TOKEN
4. Open a chat with your new bot and send /start, then send any message (e.g., "hi")
5. Visit https://api.telegram.org/bot<YOUR_TOKEN>/getUpdates in your browser
6. Find your chat ID in the JSON response:
   result[0].message.from.id — this is your TELEGRAM_CHAT_ID
7. Set both values in the CONFIG section above

NOTES:
- Keep your token secret. Anyone with it can control your bot.
- The bot can only send messages to users who have started a chat with it.
- For group chats, add the bot to the group and use the group's chat ID (usually negative).
"""


async def send_telegram_alert(
    token: str,
    chat_id: str,
    results_df: pd.DataFrame,
    market_summary: dict,
    excel_path: str,
    chart_path: str,
    candlestick_paths: list[str],
) -> None:
    """
    Send screening report to Telegram: text summary (HTML), chart photo,
    candlestick photos (media group or individually), Excel document.
    Retries on failure; truncates message if over 4096 chars.
    """
    try:
        from telegram import Bot
        from telegram.error import RetryAfter
    except ImportError:
        logger.warning("python-telegram-bot not installed. Skipping Telegram.")
        return
    bot = Bot(token=token)
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = market_summary.get("total_stocks", 0)
    n_top = min(TOP_N_STOCKS, len(results_df)) if results_df is not None and not results_df.empty else 0
    n_bot = min(5, len(results_df)) if results_df is not None and not results_df.empty else 0

    msg = f"""📊 <b>HK Stock Screening Report</b>
📅 Date: {today}
📈 Stocks Screened: {total}

━━━━━━━━━━━━━━━━━━━━━━━

📈 <b>Market Summary</b>
• Breadth (20D positive): {market_summary.get('breadth_20d_pct', 0)}%
• Above SMA50: {market_summary.get('breadth_above_sma50_pct', 0)}% | Above SMA200: {market_summary.get('breadth_above_sma200_pct', 0)}%
• GMMA: {market_summary.get('gmma_bullish', 0)} Bullish / {market_summary.get('gmma_bearish', 0)} Bearish
• Avg RSI: {market_summary.get('avg_rsi', 0)} | Overbought: {market_summary.get('rsi_overbought_count', 0)} | Oversold: {market_summary.get('rsi_oversold_count', 0)}
• Volume Spikes Today: {market_summary.get('volume_spikes_count', 0)}

━━━━━━━━━━━━━━━━━━━━━━━

🏆 <b>Top {n_top} Performers (20-Day Return)</b>
"""
    if results_df is not None and not results_df.empty:
        for rank, row in enumerate(results_df.head(n_top).to_dict("records"), 1):
            code = row.get("Stock Code", row.get("Ticker", "")).replace(".HK", "")
            ret20 = row.get("20d_return", 0) or 0
            gmma = row.get("gmma_trend", "—")
            dur = row.get("gmma_trend_duration", 0) or 0
            rsi = row.get("rsi", 0) or 0
            adx = row.get("adx", 0) or 0
            alerts = str(row.get("Alerts", ""))[:200]
            msg += f"\n{rank}. <b>{code}</b>: {ret20:+.1f}% | GMMA: {gmma} ({dur}d) | RSI: {rsi} | ADX: {adx}\n   {alerts}"
    msg += "\n\n━━━━━━━━━━━━━━━━━━━━━━━\n\n📉 <b>Bottom {n_bot} Performers</b>\n"
    if results_df is not None and not results_df.empty:
        for row in results_df.tail(n_bot).to_dict("records"):
            code = row.get("Stock Code", row.get("Ticker", "")).replace(".HK", "")
            ret20 = row.get("20d_return", 0) or 0
            gmma = row.get("gmma_trend", "—")
            dur = row.get("gmma_trend_duration", 0) or 0
            rsi = row.get("rsi", 0) or 0
            msg += f"\n<b>{code}</b>: {ret20:+.1f}% | GMMA: {gmma} ({dur}d) | RSI: {rsi}"
    msg += "\n\n━━━━━━━━━━━━━━━━━━━━━━━\n\n🔔 <b>Notable Alerts</b>\n"
    if results_df is not None and not results_df.empty and "Alerts" in results_df.columns:
        alert_rows = results_df[results_df["Alert_Count"] > 0]
        by_alert = {}
        for _, r in alert_rows.iterrows():
            for a in str(r["Alerts"]).split("|"):
                a = a.strip()
                if not a:
                    continue
                key = a.split(":")[0].strip() if ":" in a else a[:30]
                if key not in by_alert:
                    by_alert[key] = []
                code = r.get("Stock Code", r.get("Ticker", "")).replace(".HK", "")
                by_alert[key].append(code)
        for alert_type, codes in list(by_alert.items())[:15]:
            msg += f"• {alert_type}: {', '.join(codes[:8])}{' (+' + str(len(codes)-8) + ' more)' if len(codes) > 8 else ''}\n"
    if len(msg) > 4096:
        msg = msg[:4000] + "\n... and more. See Excel for full details."
    for attempt in range(TELEGRAM_RETRY_ATTEMPTS + 1):
        try:
            await bot.send_message(chat_id=chat_id, text=msg, parse_mode="HTML")
            logger.info("Telegram text message sent.")
            break
        except RetryAfter as e:
            logger.warning(f"Telegram rate limit, waiting {e.retry_after}s")
            await asyncio.sleep(e.retry_after)
        except Exception as e:
            logger.warning(f"Telegram send attempt {attempt+1} failed: {e}")
            if attempt < TELEGRAM_RETRY_ATTEMPTS:
                await asyncio.sleep(TELEGRAM_RETRY_DELAY_SECONDS)
    for attempt in range(TELEGRAM_RETRY_ATTEMPTS + 1):
        try:
            if os.path.isfile(chart_path):
                with open(chart_path, "rb") as f:
                    await bot.send_photo(chat_id=chat_id, photo=f, caption="Top performers vs HSI")
                logger.info("Telegram chart sent.")
            break
        except RetryAfter as e:
            await asyncio.sleep(e.retry_after)
        except Exception as e:
            if attempt < TELEGRAM_RETRY_ATTEMPTS:
                await asyncio.sleep(TELEGRAM_RETRY_DELAY_SECONDS)
            else:
                logger.warning(f"Telegram chart send failed: {e}")
    if candlestick_paths:
        try:
            from telegram import InputMediaPhoto
            files = []
            media = []
            for p in candlestick_paths[:10]:
                if os.path.isfile(p):
                    f = open(p, "rb")
                    files.append(f)
                    media.append(InputMediaPhoto(media=f))
            if media:
                await bot.send_media_group(chat_id=chat_id, media=media)
            for f in files:
                try:
                    f.close()
                except Exception:
                    pass
            for p in candlestick_paths[10:]:
                if os.path.isfile(p):
                    with open(p, "rb") as f:
                        await bot.send_photo(chat_id=chat_id, photo=f)
        except Exception as e:
            logger.warning(f"Telegram candlesticks send failed: {e}")
            for p in candlestick_paths:
                if os.path.isfile(p):
                    try:
                        with open(p, "rb") as f:
                            await bot.send_photo(chat_id=chat_id, photo=f)
                    except Exception:
                        pass
    for attempt in range(TELEGRAM_RETRY_ATTEMPTS + 1):
        try:
            if os.path.isfile(excel_path):
                with open(excel_path, "rb") as f:
                    await bot.send_document(chat_id=chat_id, document=f, filename=os.path.basename(excel_path))
                logger.info("Telegram Excel document sent.")
            break
        except Exception as e:
            if attempt < TELEGRAM_RETRY_ATTEMPTS:
                await asyncio.sleep(TELEGRAM_RETRY_DELAY_SECONDS)
            else:
                logger.warning(f"Telegram document send failed: {e}")


# ================================================================
# SECTION: Twilio WhatsApp Integration
# ================================================================
"""
TWILIO WHATSAPP SETUP INSTRUCTIONS:

1. Create a free Twilio account at https://www.twilio.com/try-twilio
2. Navigate to Messaging → Try it out → Send a WhatsApp message
3. Follow the sandbox setup:
   - You will see a sandbox number (e.g., +1 415 523 8886)
   - Send the activation message from your WhatsApp to that number
     (e.g., "join <two-word-code>")
   - This links your WhatsApp number to the Twilio sandbox
4. Copy your Account SID and Auth Token from the Twilio Console dashboard
5. Set the following in the CONFIG section:
   - TWILIO_ACCOUNT_SID = "your_account_sid"
   - TWILIO_AUTH_TOKEN = "your_auth_token"
   - TWILIO_WHATSAPP_FROM = "whatsapp:+14155238886"  (sandbox number)
   - TWILIO_WHATSAPP_TO = "whatsapp:+852XXXXXXXX"  (your number with country code)

NOTES:
- The sandbox expires after 72 hours of inactivity. Resend the join message to reactivate.
- For production use, apply for a WhatsApp Business API number through Twilio.
- Twilio sandbox only allows sending to numbers that have opted in (sent the join message).
- Media attachments require publicly accessible URLs.
- Free tier has limited messages per month.
"""


def send_whatsapp_alert(
    account_sid: str,
    auth_token: str,
    from_number: str,
    to_number: str,
    results_df: pd.DataFrame,
    market_summary: dict,
    excel_path: str,
    chart_path: str,
) -> None:
    """
    Send screening summary via Twilio WhatsApp (text split into multiple messages,
    chart via file.io URL if upload succeeds). Retries and rate limit (1s between messages).
    """
    try:
        from twilio.rest import Client
    except ImportError:
        logger.warning("twilio not installed. Skipping WhatsApp.")
        return
    try:
        client = Client(account_sid, auth_token)
    except Exception as e:
        logger.warning(f"Twilio client init failed: {e}")
        return
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = market_summary.get("total_stocks", 0)
    n_top = min(5, len(results_df)) if results_df is not None and not results_df.empty else 0

    msg1 = f"""📊 HK Stock Screening Report
📅 {today}
📈 Screened: {total} stocks

Market Breadth: {market_summary.get('breadth_20d_pct', 0)}% positive (20D)
GMMA: {market_summary.get('gmma_bullish', 0)} Bullish / {market_summary.get('gmma_bearish', 0)} Bearish
Volume Spikes: {market_summary.get('volume_spikes_count', 0)} stocks"""
    messages = [msg1]
    msg2 = f"🏆 Top {n_top} Performers (20D Return):\n"
    if results_df is not None and not results_df.empty:
        for i, row in enumerate(results_df.head(n_top).to_dict("records"), 1):
            code = str(row.get("Stock Code", row.get("Ticker", ""))).replace(".HK", "")
            ret = row.get("20d_return", 0) or 0
            dur = row.get("gmma_trend_duration", 0) or 0
            rsi = row.get("rsi", 0) or 0
            msg2 += f"{i}. {code}: {ret:+.1f}% | Bullish ({dur}d) | RSI {rsi}\n"
    messages.append(msg2)
    msg3 = "🔔 Key Alerts:\n"
    if results_df is not None and not results_df.empty and results_df["Alert_Count"].sum() > 0:
        alert_rows = results_df[results_df["Alert_Count"] > 0]
        by_alert = {}
        for _, r in alert_rows.iterrows():
            for a in str(r["Alerts"]).split("|"):
                a = a.strip()
                if not a:
                    continue
                key = a.split(":")[0].strip() if ":" in a else a[:25]
                if key not in by_alert:
                    by_alert[key] = []
                code = str(r.get("Stock Code", r.get("Ticker", ""))).replace(".HK", "")
                by_alert[key].append(code)
        for alert_type, codes in list(by_alert.items())[:10]:
            msg3 += f"{alert_type}: {', '.join(codes[:5])}\n"
    messages.append(msg3)
    for text in messages:
        if len(text) > 1600:
            text = text[:1580] + "..."
        for attempt in range(WHATSAPP_RETRY_ATTEMPTS + 1):
            try:
                client.messages.create(body=text, from_=from_number, to=to_number)
                logger.info("WhatsApp message sent.")
                time.sleep(1)
                break
            except Exception as e:
                logger.warning(f"WhatsApp send attempt {attempt+1}: {e}")
                if attempt < WHATSAPP_RETRY_ATTEMPTS:
                    time.sleep(WHATSAPP_RETRY_DELAY_SECONDS)
    media_url = None
    try:
        import requests
        if os.path.isfile(chart_path):
            with open(chart_path, "rb") as f:
                r = requests.post("https://file.io", files={"file": f}, timeout=30)
            if r.status_code == 200:
                data = r.json()
                media_url = data.get("link")
    except Exception as e:
        logger.debug(f"file.io upload failed: {e}")
    if media_url:
        try:
            client.messages.create(
                body="📎 Performance chart",
                from_=from_number,
                to=to_number,
                media_url=[media_url],
            )
            time.sleep(1)
        except Exception as e:
            logger.warning(f"WhatsApp media send failed: {e}")
    else:
        try:
            client.messages.create(
                body="📎 Chart and Excel report could not be attached via WhatsApp. Check Telegram or local files.",
                from_=from_number,
                to=to_number,
            )
        except Exception:
            pass


# ================================================================
# SECTION: Duplicate Report Prevention
# ================================================================

def was_report_sent_today(marker_file: str) -> bool:
    """
    Return True if marker file exists and contains today's date (YYYY-MM-DD).
    """
    if not os.path.isfile(marker_file):
        return False
    try:
        with open(marker_file, "r", encoding="utf-8") as f:
            content = f.read().strip()
        return content == datetime.now().strftime("%Y-%m-%d")
    except Exception:
        return False


def mark_report_sent(marker_file: str) -> None:
    """Write today's date (YYYY-MM-DD) to marker file."""
    try:
        with open(marker_file, "w", encoding="utf-8") as f:
            f.write(datetime.now().strftime("%Y-%m-%d"))
    except Exception as e:
        logger.warning(f"Could not write marker file: {e}")


# ================================================================
# SECTION: Optional Scheduler
# ================================================================

def run_scheduler() -> None:
    """Run main() daily at SCHEDULER_RUN_TIME if SCHEDULER_ENABLED is True."""
    try:
        import schedule
    except ImportError:
        logger.warning("schedule not installed. Run pip install schedule.")
        return
    schedule.every().day.at(SCHEDULER_RUN_TIME).do(main)
    logger.info(f"Scheduler started. Will run daily at {SCHEDULER_RUN_TIME}")
    logger.info("Press Ctrl+C to stop.")
    while True:
        schedule.run_pending()
        time.sleep(60)


# ================================================================
# SECTION: Main Execution
# ================================================================

def main() -> None:
    """End-to-end workflow: load list, download, screen, report, charts, alerts."""
    start_time = time.time()
    logger.info("=" * 60)
    logger.info("HK Stock Screening & Alert Bot")
    logger.info(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Channels: Telegram={'ON' if TELEGRAM_ENABLED else 'OFF'} | WhatsApp={'ON' if WHATSAPP_ENABLED else 'OFF'}")
    logger.info("=" * 60)

    logger.info("[Step 1/8] Loading stock list...")
    ticker_list = load_stock_list(INPUT_EXCEL_PATH)
    logger.info(f"  → Loaded {len(ticker_list)} stocks")

    logger.info("[Step 2/8] Downloading price data...")
    all_data, benchmark_data = download_all_data(ticker_list, DATA_PERIOD)
    logger.info(f"  → Successfully downloaded {len(all_data)} stocks")

    logger.info("[Step 3/8] Calculating technical indicators & screening...")
    results_df = screen_stocks(all_data)
    logger.info(f"  → Screening complete: {len(results_df)} stocks analyzed")

    logger.info("[Step 4/8] Computing market summary...")
    market_summary = compute_market_summary(results_df)
    print_market_summary(market_summary)

    logger.info("[Step 5/8] Saving Excel report...")
    save_excel_report(results_df, market_summary, OUTPUT_EXCEL_PATH)
    logger.info(f"  → Saved to {OUTPUT_EXCEL_PATH}")

    logger.info("[Step 6/8] Generating charts...")
    top_tickers = results_df.head(TOP_N_STOCKS)["Ticker"].tolist() if not results_df.empty else []
    plot_top_performers(all_data, top_tickers, benchmark_data, OUTPUT_CHART_PATH)
    candlestick_paths = plot_candlestick_charts(all_data, top_tickers, OUTPUT_CANDLESTICK_DIR)
    logger.info(f"  → Performance chart: {OUTPUT_CHART_PATH}")
    logger.info(f"  → Candlestick charts: {len(candlestick_paths)} generated")

    logger.info("[Step 7/8] Sending alerts...")
    force_send = globals().get("FORCE_SEND", False)
    if was_report_sent_today(DUPLICATE_MARKER_FILE) and not force_send:
        logger.warning("Report already sent today. Skipping. Set FORCE_SEND=True to override.")
    else:
        if TELEGRAM_ENABLED and TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID:
            logger.info("  → Sending via Telegram...")
            try:
                asyncio.run(
                    send_telegram_alert(
                        TELEGRAM_BOT_TOKEN,
                        TELEGRAM_CHAT_ID,
                        results_df,
                        market_summary,
                        OUTPUT_EXCEL_PATH,
                        OUTPUT_CHART_PATH,
                        candlestick_paths,
                    )
                )
            except Exception as e:
                logger.warning(f"Telegram send failed: {e}")
        if WHATSAPP_ENABLED and TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN:
            logger.info("  → Sending via WhatsApp...")
            try:
                send_whatsapp_alert(
                    TWILIO_ACCOUNT_SID,
                    TWILIO_AUTH_TOKEN,
                    TWILIO_WHATSAPP_FROM,
                    TWILIO_WHATSAPP_TO,
                    results_df,
                    market_summary,
                    OUTPUT_EXCEL_PATH,
                    OUTPUT_CHART_PATH,
                )
            except Exception as e:
                logger.warning(f"WhatsApp send failed: {e}")
        mark_report_sent(DUPLICATE_MARKER_FILE)
        logger.info("  → Alerts sent and marker saved.")

    if len(all_data) > 200:
        del results_df
        gc.collect()

    elapsed = time.time() - start_time
    logger.info("[Step 8/8] Complete!")
    logger.info(f"  → Total runtime: {elapsed:.1f} seconds")
    logger.info(f"  → Log file: {LOG_FILE_PATH}")
    logger.info("=" * 60)
    logger.info("✅ All done!")
    logger.info("=" * 60)


if __name__ == "__main__":
    import sys
    if "--force" in sys.argv or "-f" in sys.argv:
        globals()["FORCE_SEND"] = True
        logger.info("FORCE_SEND enabled via command line.")
    if SCHEDULER_ENABLED:
        run_scheduler()
    else:
        main()
